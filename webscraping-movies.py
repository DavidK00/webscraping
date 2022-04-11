
from cgitb import text
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

page = urlopen(webpage)			

soup = BeautifulSoup(webpage, 'html.parser')

title = soup.title

movie_table = soup.find('tbody')


movie_rows =  movie_table.findAll('tr')

#print(movie_rows[1])

for movie in range(1,6):
    td = movie_rows[movie].findAll('td')
    ranking = td[0].text
    title = td[1].text
    gross = td[5].text
    total_gross = td[7].text

    print(gross)
    input()

wb = xl.Workbook()

ws = wb.active

ws.title = 'BoxOfficeReport'

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = '% of Total Gross'


for movie in range(1,6):
    td = movie_rows[movie].findAll('td')
    ranking = td[0].text
    title = td[1].text
    gross = int(td[5].text.replace(",", "").replace('$', ""))
    total_gross = int(td[7].text.replace(",", "").replace('$', ""))
    release_date = td[8].text

    percent_gross = round(gross/total_gross * 100,2) 

    ws['A' + str(movie + 1)] = ranking
    ws['B' + str(movie + 1)] = title
    ws['C' + str(movie + 1)] = release_date
    ws['D' + str(movie + 1)] = gross
    ws['E' + str(movie + 1)] = total_gross
    ws['F' + str(movie + 1)] = str(percent_gross) + '%'

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 26

header_font = Font(size = 16, bold = True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save('BoxOfficeReport.xlsx')

